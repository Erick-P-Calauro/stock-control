from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from data import getInfo
from datetime import date

defaultTitleLine = ["Ação", "Cotação", "D.Y. (12M)", 
                  "VPA", "LPA", "Preço Teto Bazin", 
                  "Valor Intrínseco", "Margem Bazin", 
                  "Margem Graham"]

defaultTickers = ["BBAS3", "BBDC4", "SANB4"]

def generateTable(tableName="finance-table", titleContent=defaultTitleLine, tickers=defaultTickers):
    
    # Workbook and table variable
    wb = Workbook()
    table = wb.create_sheet("table", 0)

    # Row 1 height manipulation
    table.row_dimensions[1].height = 30

    # Writing title line 
    for i in range(len(titleContent)):

        # Adjusting size
        table.column_dimensions[chr(65+i)].width = 20
        
        # Reading Cell
        cel = table[chr(65 + i)+"1"]

        # Styling Cell
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.fill = PatternFill("solid", fgColor="3A9152")
        cel.font = Font(name="Roboto", size=10, color="FFFFFFFF", bold=False)

        # Settings value to cell
        cel.value = titleContent[i]

    # Writing tickers info
    for a in range(len(tickers)):
        try:
            info = getInfo(tickers[a])

            for b in range(len(info)):
                cel = table[chr(65 + b)+str(a+2)]

                cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cel.font = Font(name="Roboto", size=8)
                
                if (chr(65 + b) == "C" or chr(65 + b) == "H" or chr(65 + b) == "I"):
                    cel.number_format = "0.0%"
                elif (chr(65 + b) != "A"):
                    cel.number_format = "R$ #,##0.00"

                cel.value = info[b]
        except Exception as e:
            print("[{:s}] table build error - ".format(tickers[a]), end="")
            print(e)
    #END FOR

    # Date to table name
    today = date.today()
    day = str(today.day)
    month = str(today.month) 

    wb.save(tableName+"-"+day+"-"+month+".xlsx")